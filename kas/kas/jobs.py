# -*- coding: utf-8 -*-
import base64
import importlib
import re
import traceback
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime
from decimal import Decimal
from time import sleep

from django.conf import settings
from django.contrib.auth import get_user_model
from django.db import transaction
from django.db.models import Count, IntegerField, Q, Sum
from django.db.models.functions import Cast
from django.utils import timezone
from django.utils.translation import gettext as _
from eskat.jobs import delete_protected
from more_itertools import map_except
from openpyxl import load_workbook
from pandas import to_datetime
from prisme.models import Prisme10QBatch
from project.dafo import DatafordelerClient
from requests.exceptions import ConnectionError, HTTPError
from rq import get_current_job
from worker.job_registry import resolve_job_function
from worker.models import Job, job_decorator

from kas.eboks import EboksClient, EboksDispatchGenerator
from kas.reportgeneration.kas_final_statement import TaxFinalStatementPDF
from kas.reportgeneration.kas_report import TaxPDF
from kas.reportgeneration.kas_topdanmark_agterskrivelse import AgterskrivelsePDF

from eskat.models import (  # isort: skip
    EskatModels,
    ImportedKasMandtal,
    ImportedR75PrivatePension,
    R75SpreadsheetFile,
    get_kas_mandtal_model,
    get_r75_private_pension_model,
)
from kas.models import (  # isort: skip
    AddressFromDafo,
    Agterskrivelse,
    FinalSettlement,
    HistoryMixin,
    PensionCompany,
    PensionCompanySummaryFile,
    Person,
    PersonTaxYear,
    PersonTaxYearCensus,
    PolicyTaxYear,
    TaxSlipGenerated,
    TaxYear,
)


def mark_job_failed(job, result, exception=None):
    job.status = "failed"
    job.result = result
    if exception is not None:
        job.traceback = repr(
            traceback.format_exception(
                type(exception), exception, exception.__traceback__
            )
        )
    job.save(update_fields=["status", "result", "traceback"])
    mark_parent_job_as_failed(job)


@job_decorator
def import_mandtal(job):
    year = job.arguments["year"]
    cpr_limit = job.arguments.get("cpr")

    job.pretty_title = "%s - %s" % (job.pretty_job_type, year)

    if cpr_limit:
        job.pretty_title += " for cpr nr %s" % (cpr_limit)

    job.save()

    tax_year = TaxYear.objects.get(year=year)
    number_of_progress_segments = 2
    progress_factor = 1 / number_of_progress_segments

    mandtal_created, mandtal_updated, mandtal_cleared = ImportedKasMandtal.import_year(
        year,
        job,
        progress_factor,
        0,
        source_model=get_kas_mandtal_model(),
        cpr_limit=cpr_limit,
    )

    qs = ImportedKasMandtal.objects.filter(skatteaar=year)

    cpr_updated_list = []
    if cpr_limit is not None:
        qs = qs.filter(cpr=cpr_limit)

    count = qs.count()
    progress_start = 50
    (persons_created, persons_updated) = (0, 0)
    (persontaxyears_created, persontaxyears_updated) = (0, 0)
    (persontaxyearcensus_created, persontaxyearcensus_updated) = (0, 0)
    for i, item in enumerate(qs.iterator()):
        with transaction.atomic():
            person_data = {
                "cpr": item.cpr,
                "name": item.navn,
                "municipality_name": item.kommune,
                "municipality_code": item.kommune_no,
                "address_line_1": item.adresselinje1,
                "address_line_2": item.adresselinje2,
                "address_line_3": item.adresselinje3,
                "address_line_4": item.adresselinje4,
                "address_line_5": item.adresselinje5,
                "full_address": item.fuld_adresse,
            }

            person, status = Person.update_or_create(person_data, "cpr")
            if status == HistoryMixin.CREATED or status == HistoryMixin.UPDATED:
                cpr_updated_list.append(person_data.get("cpr"))
            if status == Person.CREATED:
                persons_created += 1
            elif status == Person.UPDATED:
                persons_updated += 1

            person_tax_year_data = {
                "person": person,
                "tax_year": tax_year,
                "number_of_days": 0,
                "fully_tax_liable": False,
            }

            person_tax_year, status = PersonTaxYear.update_or_create(
                person_tax_year_data, "tax_year", "person"
            )
            if status == PersonTaxYear.CREATED:
                persontaxyears_created += 1
            elif status == PersonTaxYear.UPDATED:
                persontaxyears_updated += 1

            person_tax_year_census_data = {
                "person_tax_year": person_tax_year,
                "imported_kas_mandtal": item.pt_census_guid,
                "number_of_days": item.skattedage,
                "fully_tax_liable": item.skatteomfang is not None
                and item.skatteomfang.lower() == "fuld skattepligtig",
            }
            person_tax_year_census, status = PersonTaxYearCensus.update_or_create(
                person_tax_year_census_data, "person_tax_year", "imported_kas_mandtal"
            )
            if status == PersonTaxYear.CREATED:
                persontaxyearcensus_created += 1
            elif status == PersonTaxYear.UPDATED:
                persontaxyearcensus_updated += 1
            person_tax_year.recalculate_mandtal(
                negative_payout_history_note=_(
                    "Genberegning på grund af import_mandtal job"
                )
            )

        if i % 1000 == 0:
            progress = progress_start + (i / count) * (100 * progress_factor)
            job.set_progress_pct(progress)

    if settings.FEATURE_FLAGS.get("enable_dafo_override_of_address"):
        dafo_client = DatafordelerClient.from_settings()
        do_mock = dafo_client.mock and not dafo_client.mock_data
        try:
            for chunk in chunks(cpr_updated_list, 100):
                requested_cprs = chunk
                if do_mock:
                    dafo_client.set_mock_data(
                        {cpr: {"cprNummer": cpr} for cpr in chunk}
                    )
                cpr_request_params = ",".join(chunk)
                params = {"cpr": cpr_request_params}
                result = dafo_client.get_person_information(params)

                for cpr in result:
                    requested_cprs.remove(cpr)
                    resultdict = result[cpr]
                    name = " ".join(
                        filter(
                            None,
                            [resultdict.get("fornavn"), resultdict.get("efternavn")],
                        )
                    )
                    address = resultdict.get("adresse", "")
                    postal_area = None
                    if resultdict.get("postnummer") and resultdict.get("bynavn"):
                        postal_area = " ".join(
                            filter(
                                None,
                                [
                                    str(resultdict.get("postnummer", 0)),
                                    resultdict.get("bynavn"),
                                ],
                            )
                        )
                    co = resultdict.get("co", "")
                    landekode = resultdict.get("landekode", "")
                    civilstand = resultdict.get("civilstand", "")
                    obj, __ = AddressFromDafo.objects.update_or_create(
                        cpr=cpr,
                        defaults={
                            "name": name,
                            "address": address,
                            "postal_area": postal_area,
                            "co": co,
                            "full_address": "\n".join(
                                filter(None, [address, postal_area])
                            ),
                        },
                    )

                    # Read the person fetched from madtal@eskat. Compare the
                    # address with the address from dafo.
                    # Overwrite the address from mandtal if the address from
                    # dafo is better
                    person = Person.objects.get(cpr=cpr)
                    person.status = "Dead" if (civilstand == "D") else "Alive"
                    if obj.is_dafo_address_better(person):
                        person.name = name
                        person.address_line_2 = address
                        person.address_line_1 = co
                        person.address_line_4 = postal_area
                        person.address_line_5 = landekode
                        person.full_address = ",".join(
                            filter(None, [address, co, postal_area, landekode])
                        )
                        person.updated_from_dafo = True

                    person.save()

                # If there is any requested cpr-numbers that is not received in
                # a response, it means that the cpr-number is invalid.
                if requested_cprs:
                    Person.objects.filter(cpr__in=requested_cprs, status="").update(
                        status="Invalid"
                    )
        finally:
            dafo_client.close()

    job.result = {
        "summary": [
            {
                "label": "Rå Mandtal-objekter",
                "value": [
                    {"label": "Tilføjet", "value": mandtal_created},
                    {"label": "Opdateret", "value": mandtal_updated},
                    {"label": "Nulstillet", "value": mandtal_cleared},
                ],
            },
            {
                "label": "Person-objekter",
                "value": [
                    {"label": "Tilføjet", "value": persons_created},
                    {"label": "Opdateret", "value": persons_updated},
                ],
            },
            {
                "label": "Personskatteår-objekter",
                "value": [
                    {"label": "Tilføjet", "value": persontaxyears_created},
                    {"label": "Opdateret", "value": persontaxyears_updated},
                ],
            },
            {
                "label": "Personskatteår-mandtal-objekter",
                "value": [
                    {"label": "Tilføjet", "value": persontaxyearcensus_created},
                    {"label": "Opdateret", "value": persontaxyearcensus_updated},
                ],
            },
        ]
    }


@job_decorator
def import_r75(job):
    year = job.arguments["year"]
    tax_year_end_date = to_datetime("%d-09-01" % (year + 1))

    job.pretty_title = "%s - %s" % (job.pretty_job_type, year)
    job.save()

    tax_year = TaxYear.objects.get(year=year)
    number_of_progress_segments = 2
    progress_factor = 1 / number_of_progress_segments
    if "source_model" in job.arguments:
        source_model = resolve_class(job.arguments["source_model"])
    else:
        source_model = get_r75_private_pension_model()

    (r75_created, r75_updated) = ImportedR75PrivatePension.import_year(
        year, job, progress_factor, 0, source_model=source_model
    )

    progress_start = 50
    (policies_created, policies_updated) = (0, 0)

    # This queryset groups results by cpr, res, ktd and creates a sum
    # of extra output field with the sum of the 'renteindtaegt' field.
    # It must be constructed with the calls in this order to generate
    # the correct SELECT .. GROUP BY .. query-
    qs = (
        ImportedR75PrivatePension.objects.values(
            "cpr", "res", "ktd", "company_pay_override"
        )
        .filter(tax_year=year)
        .annotate(
            indtaegter_sum=Sum(Cast("renteindtaegt", output_field=IntegerField()))
        )
    )
    count = qs.count()
    users_with_corrected_policies = []
    users_with_future_r75_data = []

    for i, item in enumerate(qs):
        with transaction.atomic():
            person, c = Person.objects.get_or_create(cpr=item["cpr"])
            try:
                person_tax_year = PersonTaxYear.objects.get(
                    person=person,
                    tax_year=tax_year,
                )

                res = int(item["res"])
                pension_company, c = PensionCompany.objects.get_or_create(
                    **{"res": res}
                )
                if c or pension_company.name in ("", None):
                    pension_company.name = (
                        f"Pensionsselskab med identifikationsnummer {res}"
                    )
                    pension_company.save()

                policy_data = {
                    "person_tax_year": person_tax_year,
                    "pension_company": pension_company,
                    "policy_number": item["ktd"],
                    "prefilled_amount": item["indtaegter_sum"],
                    "company_pay_override": item.get("company_pay_override", False),
                }
                (policy_tax_year, status) = PolicyTaxYear.update_or_create(
                    policy_data, "person_tax_year", "pension_company", "policy_number"
                )
                if status in (PersonTaxYear.CREATED, PersonTaxYear.UPDATED):
                    policy_tax_year._change_reason = (
                        "Updated by import"  # needed for autoligning
                    )
                    policy_tax_year.recalculate(
                        negative_payout_history_note=_(
                            "Genberegning på grund af import_r75 job"
                        )
                    )
                    policy_tax_year.save()
                    policy_tax_year._change_reason = ""
                    if status == PersonTaxYear.CREATED:
                        policies_created += 1
                    elif status == PersonTaxYear.UPDATED:
                        policies_updated += 1

                matching_r75_policies = ImportedR75PrivatePension.objects.filter(
                    cpr=item["cpr"], tax_year=year, ktd=item["ktd"], res=item["res"]
                )

                corrected = False
                r75_amounts = [int(q.renteindtaegt) for q in matching_r75_policies]

                for r75_amount in r75_amounts:
                    if -r75_amount in r75_amounts:
                        corrected = True
                        break

                if corrected:
                    # Always set corrected = True
                    person_tax_year.corrected_r75_data = True
                    person_tax_year.save(update_fields=["corrected_r75_data"])
                    users_with_corrected_policies.append(item["cpr"])
                else:
                    # Only set corrected = False if it was not set to True by this job
                    if item["cpr"] not in users_with_corrected_policies:
                        person_tax_year.corrected_r75_data = False
                        person_tax_year.save(update_fields=["corrected_r75_data"])

                r75_dates = list(
                    map_except(
                        lambda dt: to_datetime(dt, format="%Y%m%d"),
                        [str(q.r75_dato) for q in matching_r75_policies],
                        ValueError,
                    )
                )

                if len(r75_dates) > 0:
                    future_r75_data = max(r75_dates) >= tax_year_end_date
                else:
                    future_r75_data = False

                if future_r75_data:
                    # Always set future_r75_data = True
                    person_tax_year.future_r75_data = True
                    person_tax_year.save(update_fields=["future_r75_data"])
                    users_with_future_r75_data.append(item["cpr"])
                else:
                    # Only set future_r75_data = False if it was not set to
                    # True by this job.
                    if item["cpr"] not in users_with_future_r75_data:
                        person_tax_year.future_r75_data = False
                        person_tax_year.save(update_fields=["future_r75_data"])

            except PersonTaxYear.DoesNotExist:
                pass

        if i % 100 == 0:
            progress = progress_start + (i / count) * (100 * progress_factor)
            job.set_progress_pct(progress)

    job.result = {
        "summary": [
            {
                "label": "Rå R75-objekter",
                "value": [
                    {"label": "Tilføjet", "value": r75_created},
                    {"label": "Opdateret", "value": r75_updated},
                ],
            },
            {
                "label": "Policeskatteår-objekter",
                "value": [
                    {"label": "Tilføjet", "value": policies_created},
                    {"label": "Opdateret", "value": policies_updated},
                ],
            },
        ]
    }


@job_decorator
def generate_reports_for_year(job):
    qs = PersonTaxYear.get_pdf_recipients_for_year_qs(
        job.arguments["year_pk"],
        exclude_already_generated=settings.REPORT_EXCLUDE_ALREADY_GENERATED,
    )
    total_count = qs.count()
    for i, person_tax_year in enumerate(qs.iterator(), 1):
        pdf_generator = (
            TaxPDF()
        )  # construct a new pdf generator everytime to start a new pdf file
        pdf_generator.perform_complete_write_of_one_person_tax_year(
            person_tax_year, title=job.arguments["title"]
        )
        job.set_progress(i, total_count)


def chunks(lst, size):
    """Yield successive n-sized chunks from lst."""
    batch = []
    for element in lst:
        batch.append(element)
        if len(batch) >= size:
            yield batch
            batch = []
    if len(batch) > 0:
        yield batch


def update_status_for_pending_dispatches(eboks_client, pending_messages):
    for message_id_chunk in chunks(list(pending_messages.keys()), size=50):
        # Send up to 50 message_ids to get status information for
        r = eboks_client.get_recipient_status(message_id_chunk)
        for message in r.json():
            # we only use 1 recipient

            recipient = message["recipients"][0]
            if (
                message["message_id"] in pending_messages
                and recipient["post_processing_status"] != "pending"
            ):
                # if state change update it
                message = pending_messages[message["message_id"]]
                message.post_processing_status = recipient["post_processing_status"]
                message.status = "send"
                message.save(update_fields=["post_processing_status", "status"])


def mark_parent_job_as_failed(child_job, progress=None):
    parent = child_job.parent
    if parent:
        parent.status = child_job.status
        parent.result = child_job.result
        parent.traceback = child_job.traceback
        update_fields = ["status", "result", "traceback"]
        if progress:
            parent.progress = progress
            update_fields.append("progress")
        parent.save(update_fields=update_fields)


def dispatch(dispatch_qs, pending_qs, job):
    generator = EboksDispatchGenerator.from_settings()
    client = EboksClient.from_settings()
    total_count = dispatch_qs.count()
    processed = 0
    try:
        tries = 5
        while dispatch_qs.exists() and tries > 0:
            # Each iteration of the loop runs through all eligible tax slips
            # and attempts to send them. Usually the first iteration will get
            # 99% of all messages sent
            tries -= 1
            with ThreadPoolExecutor(max_workers=8) as executor:
                futures = {
                    executor.submit(
                        lambda message, client, generator: message.dispatch(
                            client, generator
                        ),
                        dispatch_item,
                        client,
                        generator,
                    ): dispatch_item
                    for dispatch_item in dispatch_qs.iterator()
                }
                for future in as_completed(futures):
                    dispatch_item = futures[future]
                    try:
                        future.result()
                    except Exception:
                        # When a message fails to send for whatever reason
                        # (usually network), skip it.
                        # It will still exist in the queryset, and subsequent tries
                        # (loop while slips.exists() and tries > 0) will attempt later
                        continue
                    if dispatch_item.status == "send":
                        processed += 1
                    job.set_progress(processed, total_count)

        if tries == 0 and dispatch_qs.exists():
            # We failed sending all messages in 5 tries
            mark_job_failed(job, "Failed sending {dispatch_qs.count()} messages")

        # Do the same loop logic again, with obtaining recipient status
        # This checks all messages that were in status "post_processing"
        # How their status is resolved in eboks.
        # Post_processing means that they are getting piped to remote printing
        # and their post_processing_status is updated to "remote printed"
        tries = 5
        first = True
        while pending_qs.exists() and tries > 0:
            tries -= 1
            if first:
                first = False
                # Remote server does take a bit of time eating through the workload
                # so give it some time to do its job
                sleep(10)

            with ThreadPoolExecutor(max_workers=8) as executor:
                futures = {}
                for chunk in chunks(pending_qs.iterator(), 10):
                    messages = {message.message_id: message for message in chunk}
                    future = executor.submit(
                        lambda message_ids: client.get_recipient_status(message_ids),
                        messages.keys(),
                    )
                    futures[future] = messages
                for future in as_completed(futures):
                    messages = futures[future]
                    try:
                        results = future.result()
                    except Exception:
                        continue
                    for result in results.json():
                        # we only use 1 recipient
                        recipient = result["recipients"][0]
                        message = messages.get(result["message_id"])
                        if (
                            message is not None
                            and recipient["post_processing_status"] != "pending"
                        ):
                            # if state change update it
                            message.post_processing_status = recipient[
                                "post_processing_status"
                            ]
                            message.status = "send"
                            message.save(
                                update_fields=["post_processing_status", "status"]
                            )
                            processed += 1
                    job.set_progress(processed, total_count)

    finally:
        client.close()

        # if we are done mark the parent job as finished
        job.result = {"dispatched_items": processed}
        job.save(update_fields=["result"])
        job.finish()


@job_decorator
def dispatch_tax_year(job):

    slips = (
        TaxSlipGenerated.objects.filter(status="created")
        .filter(
            persontaxyear__tax_year__pk=job.arguments["year_pk"],
            persontaxyear__person__is_test_person=False,
        )
        .exclude(
            persontaxyear__person__status__in=["Dead", "Invalid"],
        )
    )

    pending = TaxSlipGenerated.objects.filter(status="post_processing").filter(
        persontaxyear__tax_year__pk=job.arguments["year_pk"]
    )

    with transaction.atomic():
        job = Job.objects.select_for_update().filter(uuid=job.uuid)[0]
        job.started_at = timezone.now()
        total = slips.count()
        job.arguments["total_count"] = total
        job.arguments["current_count"] = 0
        job.save(update_fields=["status", "progress", "started_at", "arguments"])

    if total > 0:
        dispatch(slips, pending, job)
    else:
        job.finish()


def dispatch_tax_year_debug():
    rq_job = get_current_job()
    with transaction.atomic():
        job = Job.objects.select_for_update().filter(uuid=rq_job.meta["job_uuid"])[0]
        job.status = rq_job.get_status()
        job.progress = 0
        job.started_at = timezone.now()
        total = TaxSlipGenerated.objects.filter(
            Q(persontaxyear__person__name="Bent Handberg")
            | Q(persontaxyear__person__name="Person som kan logges ind på test"),
            persontaxyear__tax_year__pk=job.arguments["year_pk"],
        ).count()
        job.arguments["total_count"] = total
        job.arguments["current_count"] = 0
        job.save(update_fields=["status", "progress", "started_at", "arguments"])

    if job.arguments["total_count"] > 0:
        Job.schedule_job(
            dispatch_eboks_tax_slip_debug,
            "dispatch_tax_year_child_debug",
            job.created_by,
            parent=job,
        )
    else:
        job.finish()


@job_decorator
def dispatch_eboks_tax_slip_debug(job):
    generator = EboksDispatchGenerator.from_settings()
    slips = TaxSlipGenerated.objects.filter(
        Q(persontaxyear__person__name="Bent Handberg")
        | Q(persontaxyear__person__name="Person som kan logges ind på test"),
        persontaxyear__tax_year__pk=job.parent.arguments["year_pk"],
    )
    count = slips.count()
    client = EboksClient.from_settings()
    try:
        for slip in slips:
            message_id = client.get_message_id()
            slip.file.open(mode="rb")
            try:
                message = generator.generate_dispatch(
                    slip.title,
                    slip.persontaxyear.person.cpr,
                    pdf_data=base64.b64encode(slip.file.read()),
                )
                resp = client.send_message(
                    message=message,
                    message_id=message_id,
                )
            except (ConnectionError, HTTPError) as e:
                job.status = "failed"
                job.result = client.parse_exception(e)
                job.save(update_fields=["status", "result"])
                mark_parent_job_as_failed(job)
                break
            else:
                print(resp.json())
            finally:
                slip.file.close()
    finally:
        client.close()

    with transaction.atomic():
        parent = Job.objects.filter(pk=job.parent.pk).select_for_update()[0]
        job.set_progress(count, parent.arguments["total_count"])
        parent.result = {"dispatched_items": count}
        parent.finish()


def check_year_period(year, job, periode):
    """
    Checks that the year is in the correct periode otherwise logs the error on the job
    :param year: the year to check
    :param job: the executed job
    :param periode: the periode to check
    """
    if year.year_part != periode:
        job.status = "failed"
        job.result = {
            "error": "Kan kun {title} år som er i perioden {periode}".format(
                title=job.pretty_title, periode=periode
            )
        }
        job.end_at = timezone.now()
        job.save(update_fields=["status", "result", "end_at"])
        return False
    return True


@job_decorator
def autoligning(job):
    """Kør autoligning for et given år"""
    year = TaxYear.objects.get(pk=job.arguments["year_pk"])
    if not check_year_period(year, job, "selvangivelse"):
        return

    # if no user is found this will raise an exception which is what we want
    rest_user = get_user_model().objects.get(username="rest")

    with transaction.atomic():
        policies = (
            PolicyTaxYear.objects.select_related("person_tax_year")
            .filter(active=True, person_tax_year__tax_year=year)
            .select_for_update()
        )
        autolignet = 0
        post_processing = 0
        total = 0
        for policy in policies:
            if policy.history.filter(
                (
                    Q(
                        history_type="~",
                    )
                    & ~Q(history_change_reason="Updated by import")
                    & ~Q(history_change_reason="autoligning")
                )
                | Q(history_type="+", history_user=rest_user)
            ).exists():
                # policy has changes or policy was created by citizen
                policy.efterbehandling = True
                policy.slutlignet = False
                post_processing += 1
            elif (
                policy.person_tax_year.general_notes
                and len(policy.person_tax_year.general_notes.replace(" ", "")) > 0
            ):
                # General note on PersonTaxYear is not null and contains at
                # least 1 character
                policy.efterbehandling = True
                policy.slutlignet = False
                post_processing += 1
            elif (
                policy.person_tax_year.notes.exists()
                or policy.person_tax_year.policydocument_set.exists()
            ):
                # notes or documents was created for the related person_tax_year
                policy.efterbehandling = True
                policy.slutlignet = False
                post_processing += 1
            else:
                policy.efterbehandling = False
                policy.slutlignet = True
                autolignet += 1

            # Make sure assessed_amount is stored in the database
            policy.base_calculation_amount = policy.get_base_calculation_amount()

            # Recalculate used negative amounts etc.
            policy.recalculate(
                negative_payout_history_note=_(
                    "Genberegning på grund af autoligning job"
                )
            )

            total += 1
            policy._change_reason = "autoligning"
            policy._history_user = job.created_by
            policy.save()

        year.year_part = "ligning"
        year.save()

        job.result = {
            "autolignet": autolignet,
            "post_processing": post_processing,
            "total": total,
        }


@job_decorator
def generate_final_settlements_for_year(job):
    """
    For each person tax year generate a final settlement if:
    person_tax_year is fully tab liable and
    there exists one or more active and slutlignet policies.
    """
    tax_year = TaxYear.objects.get(pk=job.arguments["year_pk"])
    if not check_year_period(tax_year, job, "ligning"):
        return

    generated_final_settlements = 0

    qs = (
        PersonTaxYear.objects.filter(tax_year=tax_year, fully_tax_liable=True)
        .annotate(
            active_policies=Count(
                "policytaxyear",
                filter=Q(policytaxyear__active=True, policytaxyear__slutlignet=True),
            )
        )
        .filter(active_policies__gt=0)
    )
    for person_tax_year in qs.iterator():
        TaxFinalStatementPDF.generate_pdf(person_tax_year=person_tax_year)
        generated_final_settlements += 1

    job.result = {
        "status": "Genererede slutopgørelser",
        "message": generated_final_settlements,
    }


@job_decorator
def generate_batch_and_transactions_for_year(job):
    tax_year = TaxYear.objects.get(pk=job.arguments["year_pk"])
    if not check_year_period(tax_year, job, "ligning"):
        return

    collect_date = date.today().replace(month=9, day=1)
    prisme10q_batch = Prisme10QBatch(
        created_by=job.created_by, tax_year=tax_year, collect_date=collect_date
    )
    prisme10q_batch.save()

    settlements = FinalSettlement.objects.filter(
        person_tax_year__tax_year=tax_year,
        invalid=False,
        person_tax_year__person__is_test_person=False,
    )
    settlements_count = 0
    new_transactions = 0
    for final_settlement in settlements:
        settlements_count += 1
        if final_settlement.get_transaction_amount() != 0:
            prisme10q_batch.add_transaction(final_settlement)
            new_transactions += 1

    job.result = {
        "status": "Genererede batch og transaktioner",
        "message": (
            f"Genererede {new_transactions} på baggrund af"
            "{settlements_count} slutopgørelser"
        ),
    }


@job_decorator
def generate_agterskrivelser(job):
    policy_tax_year_idents = job.arguments["policy_tax_year_idents"]
    policies = [
        PolicyTaxYear.objects.filter(**policy_tax_year_ident).first()
        for policy_tax_year_ident in policy_tax_year_idents
    ]
    policies = list(filter(None, policies))

    person_tax_years_map = {}
    for policy_tax_year in policies:
        person_tax_year = policy_tax_year.person_tax_year
        if person_tax_year not in person_tax_years_map:
            person_tax_years_map[person_tax_year] = []
        person_tax_years_map[person_tax_year].append(policy_tax_year)

    job.progress = 0
    total_count = len(policies)
    job.started_at = timezone.now()
    job.save(update_fields=["status", "progress", "started_at"])

    for i, (person_tax_year, policy_tax_years) in enumerate(
        person_tax_years_map.items(), 1
    ):
        person_tax_year.agterskrivelse_set.all().delete()
        AgterskrivelsePDF.generate_pdf(person_tax_year, policy_tax_years)
        job.set_progress(i, total_count)


@job_decorator
def generate_pension_company_summary_file(job):
    pension_company = PensionCompany.objects.get(pk=job.arguments["pension_company"])
    year = TaxYear.objects.get(year=job.arguments["year"])
    PensionCompanySummaryFile.create(pension_company, year, job.created_by)

    job.result = {
        "summary": [
            {
                "label": "Årssummationsfil genereret",
                "value": [
                    {"label": "Pensionsselskab", "value": pension_company.name},
                    {"label": "Årstal", "value": year.year},
                ],
            },
        ]
    }


@job_decorator
def dispatch_final_settlements_for_year(job):
    settlements = FinalSettlement.objects.exclude(
        invalid=True, person_tax_year__person__status__in=["Dead", "Invalid"]
    ).filter(
        status="created",
        person_tax_year__tax_year__pk=job.parent.arguments["year_pk"],
        person_tax_year__person__is_test_person=False,
    )

    pending = TaxSlipGenerated.objects.filter(status="post_processing").filter(
        persontaxyear__tax_year__pk=job.arguments["year_pk"]
    )

    with transaction.atomic():
        settlements.update(title=job.arguments["title"])
        job.started_at = timezone.now()
        job.arguments["total_count"] = settlements.count()
        job.arguments["current_count"] = 0
        job.save(update_fields=["status", "progress", "started_at", "arguments"])

    if job.arguments["total_count"] > 0:
        dispatch(settlements, pending, job)

        if job.status != "failed":
            # set the current year part to genoptagelse only if job didn't fail
            year = TaxYear.objects.get(pk=job.arguments["year_pk"])
            year.year_part = "genoptagelsesperiode"
            year.save(update_fields=["year_part"])
    else:
        job.finish()


@job_decorator
def dispatch_final_settlement(job):
    """
    expects a final_settlement uuid
    :return:
    """
    client = EboksClient.from_settings()
    settlement = FinalSettlement.objects.get(uuid=job.arguments["uuid"])
    generator = EboksDispatchGenerator.from_settings()
    try:
        send_settlement = settlement.dispatch_to_eboks(client, generator)
    except Exception as e:
        mark_job_failed(job, client.parse_exception(e), e)
    else:
        if send_settlement.status == "post_processing":
            job.set_progress_pct(50)
            send_settlement.get_final_status(client)


@job_decorator
def force_finalize_settlement(job):
    """
    Forces all outstanding policies to be finalized
    :return:
    """
    policies = PolicyTaxYear.objects.filter(
        slutlignet=False, person_tax_year__tax_year__pk=job.arguments["year_pk"]
    )
    count = policies.count()
    for i, policy in enumerate(policies.iterator(), start=1):
        policy.slutlignet = True
        policy.efterbehandling = True
        policy._change_reason = "Forced finalization"
        policy._history_user = job.created_by
        policy.save()
        job.set_progress(i, count)


@job_decorator
def merge_pension_companies(job):
    try:
        target = PensionCompany.objects.get(pk=job.arguments["target"])
    except PensionCompany.DoesNotExist:
        pass
    else:
        job.set_progress_pct(50)
        with transaction.atomic():
            moved_policies = PolicyTaxYear.objects.filter(
                pension_company__in=job.arguments["to_be_merged"]
            ).update(pension_company=target)
            # This is safe and will fail, because pension_company is protected
            # on policytaxyear.
            deleted_count, _ = PensionCompany.objects.filter(
                id__in=job.arguments["to_be_merged"]
            ).delete()
        job.result = {
            "status": "Færdig",
            "message": """Flyttede %(policer)s policer
                      og flettede %(companies)s pensionsselskaber."""
            % {"policer": moved_policies, "companies": deleted_count},
        }


@job_decorator
def reset_tax_year(job):
    if settings.ENVIRONMENT not in ("development", "staging"):
        # Only allow this job to be executed in development og staging
        # the job is also hidden in the UI for production so this is just a safeguard.
        return

    with transaction.atomic():
        tax_year = TaxYear.objects.get(pk=job.arguments["year_pk"])
        TaxSlipGenerated.objects.filter(persontaxyear__tax_year=tax_year).delete()
        PensionCompanySummaryFile.objects.filter(tax_year=tax_year).delete()
        AddressFromDafo.objects.all().delete()
        delete_protected(tax_year)


@job_decorator
def generate_pseudo_settlements_and_transactions_for_legacy_years(job):
    created_final_settlements = 0
    updated_final_settlements = 0
    for tax_year in TaxYear.objects.filter(year__in=[2018, 2019]).order_by("year"):
        for person_tax_year in PersonTaxYear.objects.filter(tax_year=tax_year):
            sum_tax_after_foreign_paid_deduction = 0
            for policy in person_tax_year.policytaxyear_set.filter(active=True):
                sum_tax_after_foreign_paid_deduction += (
                    policy.full_tax - policy.foreign_paid_amount_actual
                )

            # Generate pseudo final settlement
            final_settlement, created = FinalSettlement.objects.update_or_create(
                person_tax_year=person_tax_year,
                pseudo=True,
                defaults={
                    "title": "Pseudo opgørelse",
                    "pseudo": True,
                    "pseudo_amount": sum_tax_after_foreign_paid_deduction,
                    "person_tax_year": person_tax_year,
                    "lock": person_tax_year.tax_year.get_current_lock,
                },
            )
            if created:
                created_final_settlements += 1
            else:
                updated_final_settlements += 1

    job.result = {
        "message": "Generering af pseudo slutopgørelser",
        "status": {
            "Oprettet": created_final_settlements,
            "Opdateret": updated_final_settlements,
        },
    }


@job_decorator
def import_spreadsheet_r75(job):
    file = R75SpreadsheetFile.objects.get(pk=job.arguments["pk"]).file
    company_pay_override = job.arguments["company_pay_override"]
    company = PensionCompany.objects.get(res=19625087)  # Topdanmark
    cvr = str(company.res)

    # Used to generate uuids from input data, DO NOT CHANGE!
    uuid_namespace = uuid.UUID("93874238-461b-4e7a-9989-788a3d8b46e5")

    spreadsheet_fields = {
        "Navn": (str,),
        "PoliceNr": (str, int),
        "KontoStart": (datetime,),
        "KontoOphør": (datetime,),
        "Land": (str,),
        "Adresse": (str,),
        "PensionsOrdningBeløb": (Decimal,),
        "PensionKapitalværdi": (Decimal,),
        "Kontotype": (int,),
        "År": (int,),
        "TIN Land": (str,),
        "TIN Nummer": (int,),
    }

    years = set()
    workbook = load_workbook(filename=file.path)
    policy_tax_year_idents = []
    for sheetname in workbook.sheetnames:
        sheet = workbook.get_sheet_by_name(sheetname)
        if sheet.max_row > 1:
            title_fields = []
            for rowindex, row in enumerate(sheet.iter_rows()):
                if rowindex == 0:
                    title_fields = [cell.value for cell in row]
                else:
                    rowdata = {}
                    for colindex, cell in enumerate(row):
                        value = get_formatted_cell_value(cell)
                        if colindex < len(title_fields):
                            fieldname = title_fields[colindex]
                            if type(value) is str and value.strip() == "":
                                value = None
                            if value is not None:
                                expected_type = spreadsheet_fields.get(fieldname)
                                if expected_type:
                                    if Decimal in expected_type and type(value) in (
                                        str,
                                        int,
                                        float,
                                    ):
                                        value = Decimal(value)
                                    if type(value) not in expected_type:
                                        raise Exception(
                                            f"Expected type {expected_type} in"
                                            " column {colindex+1}, "
                                            f"row {rowindex+1}, got type {type(value)} "
                                            f"for value {str(value)}"
                                        )
                            rowdata[fieldname] = value

                    if any(rowdata.values()):
                        model = EskatModels.R75SpreadsheetImport
                        year = rowdata["År"]
                        policy_number = rowdata["PoliceNr"]
                        cpr = str(rowdata["TIN Nummer"]).zfill(10)
                        years.add(year)

                        identifying_data = {
                            "ktd": policy_number,
                            "res": cvr,
                            "cpr": cpr,
                            "tax_year": year,
                        }
                        policy_tax_year_ident = {
                            "pension_company__pk": company.pk,
                            "person_tax_year__tax_year__year": year,
                            "policy_number": policy_number,
                        }
                        policy_tax_year_idents.append(policy_tax_year_ident)

                        # Get previously reported R75 amounts to override
                        if PolicyTaxYear.objects.filter(
                            **policy_tax_year_ident
                        ).exists():
                            old_imported_sum = PolicyTaxYear.objects.get(
                                **policy_tax_year_ident
                            ).prefilled_amount
                        else:
                            old_imported_sum = 0

                        model.objects.update_or_create(
                            # Identificerende; skal være den samme for en given
                            # importeret entry (ikke autogen)
                            r75_ctl_sekvens_guid=uuid.uuid5(
                                namespace=uuid_namespace, name=str(identifying_data)
                            ),
                            **identifying_data,
                            defaults={
                                "renteindtaegt": int(rowdata["PensionsOrdningBeløb"])
                                - old_imported_sum,
                                "company_pay_override": company_pay_override,
                                # Dummydata, not used in calculations
                                "pt_census_guid": uuid.uuid4(),
                                "r75_ctl_indeks_guid": uuid.uuid4(),
                                "idx_nr": 1,
                            },
                        )

    for year in years:
        Job.schedule_job(
            function=resolve_job_function("kas.jobs.import_r75"),
            job_type="ImportR75Job",
            created_by=job.created_by,
            parent=job,
            job_kwargs={
                "year": year,
                "source_model": "eskat.models:EskatModels.R75SpreadsheetImport",
            },
        )


@job_decorator
def dispatch_agterskrivelser_for_year(job):

    agterskrivelser = Agterskrivelse.objects.exclude(
        person_tax_year__person__status__in=["Dead", "Invalid"]
    ).filter(
        status="created",
        person_tax_year__tax_year__pk=job.arguments["year_pk"],
        person_tax_year__person__is_test_person=False,
    )

    pending = Agterskrivelse.objects.filter(
        status="post_processing",
        person_tax_year__tax_year__pk=job.arguments["year_pk"],
    )

    with transaction.atomic():
        job.started_at = timezone.now()
        job.arguments["total_count"] = agterskrivelser.count()
        job.arguments["current_count"] = 0
        job.save(update_fields=["status", "progress", "started_at", "arguments"])
    if job.arguments["total_count"] > 0:
        dispatch(agterskrivelser, pending, job)
    else:
        job.finish()


def resolve_class(string):
    module_string, class_name = string.rsplit(":", 1)
    item = importlib.import_module(module_string)
    for itempath in class_name.split("."):
        item = getattr(item, itempath)
    return item


number_formatters = [
    (re.compile("^(0+)$"), lambda value, match: str(value).zfill(len(match.group(1)))),
    # Add more if needed
]


def get_formatted_cell_value(cell):
    value = cell.value
    for matcher, formatter in number_formatters:
        match = matcher.fullmatch(cell.number_format)
        if match:
            return formatter(value, match)
    return value
